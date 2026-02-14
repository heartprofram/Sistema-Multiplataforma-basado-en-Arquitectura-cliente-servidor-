from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class ExportService:
    @staticmethod
    def export_stats_to_excel(stats_data, filename):
        """
        Exporta los datos de estadísticas a un archivo Excel.
        stats_data: Lista de diccionarios con las claves 'id', 'nombre', 'cargo', etc.
        filename: Ruta completa del archivo .xlsx a guardar.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Estadísticas de Asistencia"

        # --- ESTILOS ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # --- CABECERA ---
        headers = ["ID", "Empleado", "Cargo", "Días Hábiles", "Asistencias", "Efectividad %"]
        
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # --- DATOS ---
        for row_data in stats_data:
            row = [
                row_data.get("id"),
                row_data.get("nombre"),
                row_data.get("cargo"),
                row_data.get("dias_habiles"),
                row_data.get("asistencias"),
                f"{row_data.get('efectividad')}%"
            ]
            ws.append(row)
            
            # Aplicar bordes a la nueva fila
            current_row = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = thin_border
                cell.alignment = center_align

        # --- AUTO-AJUSTAR ANCHO DE COLUMNAS ---
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        try:
            wb.save(filename)
            return True, "Archivo guardado correctamente."
        except Exception as e:
            return False, f"Error al guardar: {e}"
